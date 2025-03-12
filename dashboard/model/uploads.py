from functools import cached_property
from flask import current_app
from lbrc_flask.database import db
from lbrc_flask.security import AuditMixin
from lbrc_flask.model import CommonMixin
from lbrc_flask.validators import is_integer, parse_date_or_none
from sqlalchemy.orm import Mapped, mapped_column
from sqlalchemy import String, Text
from werkzeug.utils import secure_filename
from itertools import takewhile
from openpyxl import load_workbook


class Upload(AuditMixin, CommonMixin, db.Model):
    COLUMNS = {
        'Project Title': dict(type='str'),
        'Local REC number': dict(type='str', max_length=20),
        'Is this project sensitive?': dict(type='yn'),
        'Project summary': dict(type='str'),
        'IRAS Number': dict(type='int', allow_nulls=True),
        'CRN/RDN Portfolio study': dict(type='yn'),
        'CRN/RDN CPMS ID': dict(type='int'),
        'First Name': dict(type='str', max_length=100),
        'Surname': dict(type='str', max_length=100),
        'ORCID': dict(type='str', max_length=100),
        'Project Actual Start Date': dict(type='date'),
        'Project End Date ': dict(type='date', allow_nulls=True),
        'Project Status': dict(type='str', max_length=100),
        'Theme': dict(type='str', max_length=100),
        'UKCRC Health Category': dict(type='str', max_length=100),
        'NIHR priority Areas / Fields of Research': dict(type='str', max_length=100),
        'REC Approval Required': dict(type='yn'),
        'UKCRC Research Activity Code': dict(type='str', max_length=100),
        'RACS sub-categories': dict(type='str', max_length=100, allow_null=True),
        'Research Type': dict(type='str', max_length=100),
        'Methodology': dict(type='str', max_length=100),
        'Expected Impact': dict(type='str', max_length=100),
        'Randomised Trial': dict(type='yn'),
        'Trial Phase': dict(type='str', max_length=100, allow_null=True),
        'Participants Recruited to Centre FY': dict(type='int'),
        'BRC funding ': dict(type='int'),
        'Main Funding Source': dict(type='str', max_length=100),
        'Main Funding Category': dict(type='str', max_length=100),
        'Main Funding - BRC Funding': dict(type='str', max_length=100),
        'Main Funding - DHSC/NIHR Funding': dict(type='str', max_length=100),
        'Main Funding - Industry Collaborative or Industry Contract Funding': dict(type='str', max_length=100),
        'Total External Funding Awarded': dict(type='int'),
        'First in Human Project': dict(type='yn'),
        'Link to NIHR Translational Research Collaboration': dict(type='yn'),
        'Comments': dict(type='str'),
    }

    STATUS__AWAITING_PROCESSING = 'Awaiting Processing'
    STATUS__PROCESSED = 'Processed'
    STATUS__ERROR = 'Error'

    STATUS_NAMES = [
        STATUS__AWAITING_PROCESSING,
        STATUS__PROCESSED,
        STATUS__ERROR,
    ]

    COLUMN_NAMES = set(COLUMNS.keys())

    id: Mapped[int] = mapped_column(primary_key=True)
    filename: Mapped[str] = mapped_column(String(500))
    status: Mapped[str] = mapped_column(String(50), default='')
    errors: Mapped[str] = mapped_column(Text, default='')

    @property
    def local_filepath(self):
        return current_app.config["FILE_UPLOAD_DIRECTORY"] / secure_filename(f"{self.id}_{self.filename}")

    def worksheet(self):
        wb = load_workbook(filename=self.local_filepath, read_only=True)
        return wb.active

    @cached_property
    def column_names(self):
        rows = self.worksheet().iter_rows(min_row=1, max_row=1)
        first_row = next(rows)

        result = [c.value.lower() for c in takewhile(lambda x: x.value, first_row)]

        return result

    def iter_rows(self):
        for r in self.worksheet().iter_rows(values_only=True):
            values = dict(zip(self.column_names, r))
            yield values

    def iter_data(self):
        for r in self.iter_rows:
            if self._is_data_row(r):
                yield r

    def _is_data_row(self, row):
        return 'key' in row.keys() and (row.get('key', None) is None or is_integer(row['key']))

    def validate(self):
        errors = []

        errors.extend(self._column_validation_errors())
        errors.extend(self._data_validation_errors())

        if errors:
            self.errors = "\n".join(errors)
            self.status = Upload.STATUS__ERROR
    
    def _column_validation_errors(self):
        missing_columns = Upload.COLUMN_NAMES - set(self.column_names)
        return map(lambda x: f"Missing column '{x}'", missing_columns)

    def _data_validation_errors(self):
        errors = []

        for i, row in enumerate(self.iter_rows):
            if self._is_ambigous_row(row):
                errors.append(f"Row {i}: contains columns for both bacteria and phages")
            elif self._neither_phage_nor_bacterium(row):
                errors.append(f"Row {i}: does not contain enough information")
            elif errors := self._bacterium_errors(row):
                for e in errors:
                    errors.append(f"Row {i}: {e}")
            elif errors := self._phage_errors(row):
                for e in errors:
                    errors.append(f"Row {i}: {e}")

        return errors

    def _is_invalid_string(self, value, column, col_def):
        max_length = col_def.get('max_length', None)
        
        if not max_length:
            return
        
        if value is None or len(value) > max_length:
            return f"Text too long in column '{column}'"

    def _is_invalid_interger(self, value, column, col_def):
        if value is None or not is_integer(value):
            return f"Invalid value in column '{column}'"

    def _is_invalid_date(self, value, column, col_def):
        if parse_date_or_none(value) is None:
            return f"Invalid value in column '{column}'"

    @property
    def is_error(self):
        return self.status == Upload.STATUS__ERROR
